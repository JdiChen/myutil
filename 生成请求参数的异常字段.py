import copy
import json
import os.path
import re
import sys
import time

from loguru import logger as log
from openpyxl import load_workbook
from faker import Faker
from openpyxl.worksheet.worksheet import Worksheet

from base import BaseApi


class GenerateBody:
    """
    bug:当字段值为列表时,也是赋值的"" 需要改成[]
    """

    def __init__(self, source_body, config_body):
        self.source_body = source_body
        self.config_body = config_body
        self.gen_conf = self.__get_generate_conf()

    def __is_config_true(self, config_key: str) -> bool:
        """
        配置文件中有的，就返回True
        :param config_key: 配置文件中的key名称
        :return:
        """
        tmp = ["maxLen", "minNum", "maxNum", "update", "required"]
        if config_key in tmp:
            if self.gen_conf[config_key] == 0:
                return True
        else:
            if self.gen_conf["required"][config_key] == 0:
                return True
        return False

    def __gen_key(self, gen_key: str, key, tmp, val=None):
        """
        生成body
        :param gen_key:
        :param key:
        :param tmp:
        :param val:
        :return:
        """
        if gen_key == "delKey":
            return pop_val(key, "tmp", tmp)
        return set_val(key, "tmp", tmp, val)

    def __get_val(self, required_key, key, val, desc):
        fake = Faker("zh-CN")
        if required_key == "maxLen":
            lenght = self.config_body[key]["maxLen"]
            val = "".join(fake.random_letters(lenght + 1))
            desc = f"长度为{lenght}+1"
        elif required_key == "min":
            lenght = self.config_body[key]["min"]
            val = int(lenght) - 1
            desc = f"最小值为{lenght}-1"
        elif required_key == "max":
            lenght = self.config_body[key]["max"]
            val = int(lenght) + 1
            desc = f"最大值为{lenght}+1"
        return desc, val

    def __key_genrate(self, config_key: str, required_key: str, desc: str = None, val: str = None) -> list:
        """

        :param required_key: 传入配置文件内对应的key
        :param desc: 最终生成的描述key
        :param val:  需要给生成key的值
        :return:
        """
        required_bodys = []
        # 配置为0表示不需要检验,直接返回空列表
        if self.__is_config_true(config_key):
            return required_bodys
            # 字段循环
        for key in self.config_body.keys():
            if required_key not in self.config_body[key]:
                continue
            # 原参数为列表则取第一个值来使用，后面再将其转为列表
            desc, val = self.__get_val(required_key, key, val, desc)
            if type(self.source_body) is list:
                tmp = copy.deepcopy(self.source_body[0])
                required_bodys.append({f"{key}{desc}": [self.__gen_key(config_key, key, tmp, val)]})
            else:
                tmp = copy.deepcopy(self.source_body)
                required_bodys.append({f"{key}{desc}": self.__gen_key(config_key, key, tmp, val)})
        log.info(f"生成完成,生成个数为{len(required_bodys)}")
        log.info(json.dumps(required_bodys))
        return required_bodys

    @staticmethod
    def __get_generate_conf():
        path = "config.json"
        with open(path, "r", encoding='utf8') as f:
            conf = json.load(f)
        # log.debug(conf)
        return conf

    def __required_delKey_args_generate(self) -> list:
        log.debug("开始生成必填字段删除字段")
        required_bodys = self.__key_genrate("delKey", "required", "删除")
        return required_bodys

    def __required_empty_args_generate(self) -> list:
        """
        需要生成 不传键值,  值为""   值为 " "
        :return:
        """
        log.debug("开始生成必填字段为空")
        required_bodys = self.__key_genrate("empty", "required", "为空", "")
        return required_bodys

    def __required_space_args_generate(self) -> list:
        """
        需要生成 不传键值,  值为""   值为 " "
        :return:
        """
        log.debug("开始生成必填字段为空格")
        required_bodys = self.__key_genrate("space", "required", "为空格", " ")
        return required_bodys

    def __max_len_args_generate(self) -> list:
        """
        超长检验
        :return:
        """
        log.debug("开始生成字段最大字符串加1长度")
        max_len_bodys = self.__key_genrate("maxLen", "maxLen")
        return max_len_bodys

    def __min_number_args_generate(self, ):
        log.debug("开始生成字段最小值 -1 ")
        min_number_bodys = self.__key_genrate("minNum", "min")
        return min_number_bodys

    def __max_number_args_generate(self, ):
        log.debug("开始生成字段最大值 +1 ")
        max_number_bodys = self.__key_genrate("maxNum", "max")
        return max_number_bodys

    def __update_args_generate(self, ):
        return []

    @staticmethod
    def megre_bodys(*args) -> list:
        req_bodys = []
        for arg in args:
            req_bodys += arg
        log.info("合并成功")
        log.debug(req_bodys)
        log.debug(len(req_bodys))
        return req_bodys

    def generate(self):
        empty = self.__required_empty_args_generate()
        space = self.__required_space_args_generate()
        del_key = self.__required_delKey_args_generate()
        max_len = self.__max_len_args_generate()
        min_num = self.__min_number_args_generate()
        max_num = self.__max_number_args_generate()
        update = self.__update_args_generate()
        return self.megre_bodys(empty, space, del_key, max_len, min_num, max_num, update)


def get_val(body: dict, text: str, val):
    """
     k1.k2[2].k3 方式   dict[k1][k2][2][k3] 的方式 获取值
    """
    keys = text.split(".")
    for key in keys:
        index = get_index(key)
        log.debug(f"index is {index}")
        log.debug(f"key is {key}")
        if index != -1:
            t = key.split("[")[0]
            log.debug(f"t is {t}")
            log.debug(f"get body[{t}][{index}]")
            body = body[t][index]
        else:
            log.debug(f"get body[{key}]")
            body = body[key]
        log.debug(f"tmp is {body}")
    log.debug(f"tmp is {body}")


def set_val(text: str, body_name: str, req_json: dict, val) -> dict:
    """
    将字典查找的字符串转换为 字符串执行命令
    exec("body['snList'][0]['defectCode'][0]['must']=1")
    eval("body['snList'][0]['defectCode'][0]['must']")
       """
    keys = text.split(".")
    exec_str = f"{body_name}"
    for key in keys:
        if "[" in key:
            index = get_index(key)
            t = key.split('[')[0]
            exec_str += f"[\"{t}\"][{index}]"
        else:
            exec_str += f"[\"{key}\"]"
    if isinstance(val, int):
        exec_str += f"={val}"
    else:
        exec_str += f"=\"{val}\""
    # print(exec_str)
    # exec内的变量需要以字典传入  {变量名:变量值},否则读取的是全局变量
    exec(exec_str, {body_name: req_json})
    return req_json


def pop_val(text: str, body_name, body) -> dict:
    keys = text.split(".")
    exec_str = f"del {body_name}"
    for key in keys:
        if "[" in key:
            index = get_index(key)
            t = key.split('[')[0]
            exec_str += f"[\"{t}\"][{index}]"
        else:
            exec_str += f"[\"{key}\"]"
    # log.debug(exec_str)
    # exec内的变量需要以字典传入  {变量名:变量值},否则读取的是全局变量
    exec(exec_str, {body_name: body})
    # log.debug(body)
    return body


def get_index(text: str) -> int:
    """
    获取 k1[5] 样式中的 5
    :param text:
    :return:
    """
    find_list = re.findall("\d+", text)
    if find_list and "[" in text:
        return int(find_list[-1])
    return -1


def add_key(body, key, val):
    """
    从ecxel读取配置时,如果读到空,就不加入配置json内
    如果值不为空就添加进字典
    :param body:
    :param key:
    :param val:
    :return:
    """
    if val is not None:
        body[key] = val


def read_excel(path) -> dict:
    """
    从文件内读取配置,第一列第二行开始为字典的键,
    每一行后面再组成一个字典作为行首键的值
    {
        key:
            {
            key:va1
            }
    }
    :return:
    """
    log.info(f"开始读取excel,路径为{os.path.abspath(path)}")
    wb = load_workbook(path)
    conf_sheet: Worksheet = wb[wb.sheetnames[0]]
    url_sheet: Worksheet = wb[wb.sheetnames[1]]
    log.info(f"读取{conf_sheet} {url_sheet} sheet页")
    rows = conf_sheet.max_row  # 行
    # 获取配置
    req_json = {}
    # 这里openpyxl模块按ecxel的行号做为下标开始读的
    for i in range(2, rows + 1):
        tmp = {}
        j = 1
        add_key(tmp, "required", conf_sheet[i][j].value)
        add_key(tmp, "maxLen", conf_sheet[i][j + 1].value)
        add_key(tmp, "min", conf_sheet[i][j + 2].value)
        add_key(tmp, "max", conf_sheet[i][j + 3].value)
        add_key(tmp, "canUpdate", conf_sheet[i][j + 4].value)
        req_json[conf_sheet[i][0].value] = tmp
    log.info("配置获取完成")
    log.info(f"{req_json}")

    # 获取请求信息
    url = url_sheet["A2"].value
    # 这里读取出来为字符串,调用时需要用json.loads转为字典
    source_body: str = url_sheet["B2"].value
    method: str = url_sheet["C2"].value
    log.info("请求信息获取完成")
    log.info(f"url={url}")
    log.info(f"method={method}")
    log.info(f"body={source_body}")

    info = {"url": url, "source_body": source_body, "method": method.upper(), "req_json": req_json}
    return info


def main(args):
    host = r"you host url"
    path = r"%s" % args[-1]
    if not path.endswith(".xlsx"):
        log.error("请使用xlsx文件")
        sys.exit(-1)
    info = read_excel(path)
    # log.add(info["url"]+get_time(), encoding="utf8", backtrace=True, mode="w", rotation="100 MB")
    body = json.loads(info["source_body"])
    requirement_json = info["req_json"]
    gb = GenerateBody(body, requirement_json)
    body_list = gb.generate()
    send = BaseApi()
    # 需要造数据的怎么处理？
    for req in body_list:
        items = req.items()
        for key, val in items:
            log.debug(
                f"======================================== 验证 {key} =================================================")
            send.req(host + info['url'], info['method'], val)


def get_time():
    if not os.path.exists("log"):
        os.mkdir("log")
    fmt = "%Y%m%d_%H%M%S.txt"
    filename = time.strftime(fmt, time.localtime())
    return os.path.join("log", filename)


if __name__ == '__main__':
    # file_name = r"%s.txt" % get_time()
    log.add(os.path.join(get_time()), encoding="utf8", backtrace=True, mode="w", rotation="100 MB")
    main(sys.argv)
