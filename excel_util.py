from openpyxl import load_workbook


class ExcelRW:
    work_book = None

    def __init__(self, path: str):
        if not path.endswith('.xlsx'):
            print("plese use xlsx file")
            raise FileNotFoundError
        self.work_book = load_workbook(path,read_only=True)

    def __del__(self):
        self.work_book.close()

    def get_all_sheet(self):
        return self.work_book.sheetnames

    def get_sheet_row_col(self, sheet_name):
        # return self.work_book[sheet_name].active_cell #当前定位的单元格
        """
        :param sheet_name:
        :return: （有数据的总行数，总列数）
        """
        sheet = self.work_book[sheet_name]
        return sheet.max_row, sheet.max_column

    def get_all_cell_value(self, sheet_name):
        """
        :param sheet_name:
        :return:all test cases list
        """
        cell_list = []
        row, col = self.get_sheet_row_col(sheet_name)
        sheet = self.work_book[sheet_name]
        for r in range(1,row+1):
            tmp_dict = {}
            for c in range(1,col+1):
                cell_value = sheet.cell(r, c).value
                first_col_value = sheet.cell(1,c).value
                tmp_dict[first_col_value] = cell_value
            cell_list.append(tmp_dict)
        return cell_list

    def cell_to_dict(self):
        suite_dict = {
            "test_name": None,
            "setup": None,
            "steps": [],
            "asser": None,
            "teardown": None
        }
        cases = ExcelRW(path).get_all_cell_value('test_contacts')
        for i in range(len(cases)):
            suite_dict["test_name"] = cases[i+1]['test_name']
            suite_dict["steps"].append(cases[i+1]['step_name'])
            suite_dict["steps"].append(cases[i+1]['location_method'])
            suite_dict["steps"].append(cases[i+1]['location_str'])
            suite_dict["steps"].append(cases[i+1]['action'])

if __name__ == '__main__':
    # path = 'test.xlsx'
    path = r'E:\PythonProject\ui_test_by_keyword\testcase\6801_testcase.xlsx'
    a = ExcelRW(path).get_all_cell_value('test_contacts')
    for i in a:
        print(i)